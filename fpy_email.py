#!/usr/bin/env python
# -*- coding: utf-8 -*-
import smtplib
import openpyxl
import datetime
import os
import prettytable
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import load_workbook

# Get current date
today = datetime.datetime.today().strftime('%Y-%m-%d')

# get yesterday's date as a datetime object
yesterday = datetime.datetime.today() - datetime.timedelta(days=1)

# format yesterday's date as a string in the format YYYY-MM-DD
yesterday_str = yesterday.strftime('%Y-%m-%d')

# set the desired filename
filename = yesterday_str + '_yield.xlsx'

# Load workbook
wb = load_workbook(filename=filename, read_only=True)

# Select worksheet
ws = wb['Summary']


# Create message
msg = MIMEMultipart()
msg['Subject'] = '{} - EVE SLT'.format(yesterday_str)
msg['From'] = 'samson.lau2@hyvesolutions.com'
#msg['To'] = 'samson.lau2@hyvesolutions.com'
to_emails = ['samson.lau2@hyvesolutions.com', 'samuel.cheung@hyvesolutions.com', 'yaochang.he@hyvesolutions.com', 'ravneet.chohan@hyvesolutions.com', 'austin.lin@hyvesolutions.com' , 'ravip@hyvesolutions.com' , 'stephen.woodbury@hyvesolutions.com' , 'jitendras@hyvesolutions.com' , 'kennethhu@hyvesolutions.com' , 'DongqiL@hyvesolutions.com']
#to_emails = ['samson.lau2@hyvesolutions.com', 'yaochang.he@hyvesolutions.com']
msg['To'] = ', '.join(to_emails)

# set the desired filename
filename = yesterday_str + '_yield.xlsx'

# Add attachment to the message
#attachment = open(filename, 'rb')
#part = MIMEApplication(attachment.read(), Name=os.path.basename(filename))
#part['Content-Disposition'] = 'attachment; filename="{filename}"'
#msg.attach(part)

# Add attachment to the message
with open(filename, 'rb') as attachment:
    part = MIMEApplication(attachment.read(), Name=os.path.basename(filename))
    part['Content-Disposition'] = 'attachment; filename="%s"' % os.path.basename(filename)
    msg.attach(part)


# Read Summary Sheet content from the file
wb = openpyxl.load_workbook(filename, read_only=True)

# Function to extract cell values from a given sheet and range
def get_sheet_values(sheet, min_row, max_row, min_col, max_col):
    cell_values = []
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        cell_values.append([cell.value for cell in row])
    return cell_values

# Add content from the "Summary" sheet to email body
ws_summary = wb['Summary']
summary_values = get_sheet_values(ws_summary, 1, 4, 1, 5)
email_body = yesterday_str + ' Yield (EVE):\n'
for row in summary_values:
    email_body += '\t'.join([str(cell) for cell in row]) + '\n'

# Add content from the "Pivot Yield" sheet to email body
ws_pivot_yield = wb['Pivot Yield']
# Determine the range dynamically based on the content of the sheet
max_row = ws_pivot_yield.max_row
max_col = ws_pivot_yield.max_column
pivot_yield_values = get_sheet_values(ws_pivot_yield, 5, max_row, 1, max_col)
email_body += '\nPivot Table: Number of PASS/ FAIL sn_tag by Model:\n'
for row in pivot_yield_values:
    email_body += '\t'.join([str(cell) for cell in row]) + '\n'

# Attach the email body content to the message
msg.attach(MIMEText(email_body))


# Create SMTP object
smtp_server = 'fca-savsmtp.synnex.org'
smtp_port = 25
smtp = smtplib.SMTP(smtp_server, smtp_port)

# Send email
try:
    smtp.sendmail(msg['From'], to_emails, msg.as_string())
    print('Email sent successfully')
except smtplib.SMTPException as e:
    print('Failed to send email. Error:', e)

# Close SMTP connection
smtp.quit()















"""
# Read Summary Sheet content from the file
wb = openpyxl.load_workbook(filename, read_only=True)
ws = wb['Summary']
cell_values = []
for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=5):
    cell_values.append([cell.value for cell in row])

# Add Summary Sheet content to email body
email_body = yesterday_str + ' Yield (EVE):\n'
for row in cell_values:
    email_body += '\t'.join([str(cell) for cell in row]) + '\n'
msg.attach(MIMEText(email_body))

# Create SMTP object
smtp_server = 'fca-savsmtp.synnex.org'
smtp_port = 25
smtp = smtplib.SMTP(smtp_server, smtp_port)

# Send email
try:
    smtp.sendmail(msg['From'], to_emails, msg.as_string())
    print('Email sent successfully')
except smtplib.SMTPException as e:
    print('Failed to send email. Error:', e)

# Close SMTP connection
smtp.quit()

"""